# file: extract_clauses.py

from fastapi import UploadFile
import json
import shutil
import requests
from openpyxl import load_workbook, Workbook
from docx import Document

# ✨ إعداد كلاس LLM
class ChatBotHelper:
    def __init__(self, excel_examples, word_text):
        self.memory = []
        self.url = "https://qwen-vlm-a100.delightfulsky-f308bdb7.westus3.azurecontainerapps.io/qwen/generate_pdf"
        self.system_message = (
            "📚 You are an intelligent assistant that can extract structured regulatory or contractual clauses (Arabic or English) "
            "from messy, unstructured Word documents.\n\n"
            "You will be given:\n"
            "1️⃣ Examples of clean, structured clauses extracted from Excel.\n"
            "2️⃣ Raw text from a Word file containing messy or free-form legal or policy content.\n\n"
            "Your job:\n"
            "- Learn the structure and formatting from the Excel examples.\n"
            "- Read the Word content carefully.\n"
            "- Extract all important clauses and format them the same way.\n\n"
            "Return the result in the following **JSON structure**:\n"
            "{\n  \"clauses\": [\n    { \"title\": \"البند: 1\", \"description\": \"التوضيح: ...\" }, ... ]\n}\n\n"
            "==================================\n"
            "📘 Excel Examples:\n" + excel_examples +
            "\n==================================\n"
            "📄 Word Content:\n" + word_text +
            "\n==================================\n"
            "Now extract the clauses from the Word content and return the JSON only.\n"
        )
        self.init_conversation()

    def init_conversation(self):
        self.memory.append({"role": "system", "content": self.system_message})

    def chat(self, user_message):
        self.memory.append({"role": "user", "content": user_message})
        messages = self.memory
        json_payload = json.dumps([m["content"] for m in messages])

        try:
            response = requests.post(
                self.url,
                data={
                    'messages': json_payload,
                    'max_new_tokens': "3000",
                }
            )
            return response.text if response.status_code == 200 else f"Error: {response.status_code}"
        except Exception as e:
            return f"Exception: {str(e)}"


def extract_text_from_docx(path):
    doc = Document(path)
    return "\n".join([p.text.strip() for p in doc.paragraphs if p.text.strip()])


def extract_examples_from_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    return "\n\n".join(
        [f"- البند: {r[0]}\n  التوضيح: {r[1]}" for r in ws.iter_rows(min_row=2, values_only=True) if r[0] and r[1]]
    )


def save_clauses_to_excel(clauses, output_excel_path="./database/output_clauses.xlsx"):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Clauses"

    ws.append(["رقم البند", "الوصف"])  # Header row

    for clause in clauses:
        title = clause.get("title", "").replace("البند:", "").strip()
        description = clause.get("description", "").replace("التوضيح:", "").strip()
        if title or description:  # Avoid empty rows
            ws.append([title, description])

    wb.save(output_excel_path)


def extract_json_from_response(raw_response):
    try:
        outer = json.loads(raw_response) if isinstance(raw_response, str) else raw_response
        inner = outer.get("response", "")

        # إذا كانت الاستجابة تحتوي على ```json
        if isinstance(inner, str) and "```json" in inner:
            json_start = inner.find("```json") + len("```json")
            json_end = inner.find("```", json_start)
            clean_json = inner[json_start:json_end].strip()
            return json.loads(clean_json)

        # لو الاستجابة مباشرة JSON
        elif isinstance(inner, dict) and "clauses" in inner:
            return inner

        # لو الاستجابة نص JSON بدون العلامات
        elif isinstance(inner, str):
            return json.loads(inner.strip())

        else:
            print("⚠️ Unexpected format in response")
            return None

    except Exception as e:
        print("❌ Failed to parse JSON from response:", e)
        return None


def save_temp_file(uploaded_file: UploadFile, save_path: str):
    with open(save_path, "wb") as buffer:
        shutil.copyfileobj(uploaded_file.file, buffer)
    return save_path

# ========================================
# 🔧 MAIN EXECUTION
# if __name__ == "__main__":
 
#     word_file = "H:\\GlassHub\\cst_agent\\database\\glasshub_files\\تنظيمات قواعد حماية حقوق المستخدم.docx"
#     excel_file = "C:\\Users\\Mohamed Hamde\\Downloads\\بنود مصفوفة الإلتزام لعدد من التنظيمات.xlsx"

#     url = "https://qwen-vlm-a100.delightfulsky-f308bdb7.westus3.azurecontainerapps.io/qwen/generate_pdf"

   
#     word_text = extract_text_from_docx(word_file)
#     excel_examples = extract_examples_from_excel(excel_file)

    
#     bot = ChatBotHelper(excel_examples, word_text)

#     print("🚀 Sending request to LLM...")
#     response = bot.chat("استخرج البنود الهامة من النص الموجود في ملف الوورد بنفس شكل الأمثلة في الإكسل.")
#     with open("raw_response.txt", "w", encoding="utf-8") as debug_file:
#         debug_file.write(repr(response))


#     print("\n📦 Result:")
#     print(repr(response))


#     try:
#         parsed = extract_json_from_response(response)

#         if parsed:

#             with open("./database/output_clauses.json", "w", encoding="utf-8") as f:
#                 json.dump(parsed, f, ensure_ascii=False, indent=2)
#             print("✅ Saved to output_clauses.json")


#             clauses = parsed.get("clauses", [])
#             if isinstance(clauses, list) and len(clauses) > 0:
#                 save_clauses_to_excel(clauses)
#             else:
#                 print("⚠️ No clauses found in parsed JSON.")
#         else:
#             print("⚠️ Could not extract JSON from response.")

#     except Exception as e:
#         print("⚠️ Could not save JSON:", e)

